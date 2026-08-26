"""خدمات التصدير — تصدير بيانات إلى Excel."""

import io

from openpyxl import Workbook

from app.core.i18n import _
from app.models.class_room import ClassMember
from app.models.gradebook import GradeEntry, GradeItem
from app.models.progress import StudentProgress


def export_students_excel(class_id: int) -> bytes:
    """تصدير قائمة الطلاب إلى Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = _("الطلاب")
    ws.append([_("الاسم"), _("البريد الإلكتروني"), _("تاريخ الانضمام"), _("الحالة")])

    members = ClassMember.query.filter_by(class_id=class_id).all()
    for m in members:
        ws.append(
            [
                m.user.name_ar or "",
                m.user.email,
                m.joined_at.strftime("%Y-%m-%d") if m.joined_at else "",
                _("نشط") if m.status == "active" else m.status,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_grades_excel(class_id: int) -> bytes:
    """تصدير درجات الصف إلى Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "الدرجات"

    items = GradeItem.query.filter_by(class_id=class_id).order_by(GradeItem.id).all()
    header = ["الاسم", "البريد الإلكتروني"]
    for item in items:
        header.append(f"{item.title} (/{item.max_mark or '?'})")
    ws.append(header)

    members = ClassMember.query.filter_by(class_id=class_id, status="active").all()
    entries = {}
    if items:
        rows = GradeEntry.query.filter(GradeEntry.grade_item_id.in_([i.id for i in items])).all()
        entries = {(r.student_id, r.grade_item_id): r for r in rows}

    for m in members:
        row = [m.user.name_ar or "", m.user.email]
        for item in items:
            entry = entries.get((m.user_id, item.id))
            row.append(float(entry.mark) if entry and entry.mark is not None else "")
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_moe_format(school_id=None, academic_year=None, term=None):
    from app.extensions import db
    from app.models.class_room import ClassMember, ClassRoom
    from app.models.gradebook import GradeEntry, GradeItem
    from app.models.school import Grade, School

    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات الوزارة"
    ws.append(
        [
            "student_id",
            "student_name",
            "grade_level",
            "subject",
            "final_score",
            "academic_year",
            "school_name",
        ]
    )

    query = ClassRoom.query.filter_by(is_active=True)
    if school_id:
        query = query.filter_by(school_id=school_id)
    classes = query.all()
    for cls in classes:
        items = GradeItem.query.filter_by(class_id=cls.id).all()
        item_ids = [i.id for i in items]
        if not item_ids:
            continue
        entries = GradeEntry.query.filter(GradeEntry.grade_item_id.in_(item_ids)).all()
        student_marks: dict[int, list[float]] = {}
        for e in entries:
            if e.student_id not in student_marks:
                student_marks[e.student_id] = []
            if e.mark is not None:
                student_marks[e.student_id].append(float(e.mark))
        members = ClassMember.query.filter_by(class_id=cls.id, status="active").all()
        for m in members:
            marks = student_marks.get(m.user_id, [])
            avg_score = round(sum(marks) / len(marks), 2) if marks else 0
            grade_obj = db.session.get(Grade, cls.grade_id) if cls.grade_id else None
            school_obj = db.session.get(School, cls.school_id) if cls.school_id else None
            ws.append(
                [
                    m.user_id,
                    m.user.name_ar or m.user.email,
                    grade_obj.grade_level if grade_obj else "",
                    cls.subject.name_ar if cls.subject else "",
                    avg_score,
                    academic_year or "",
                    school_obj.name_ar if school_obj else "",
                ]
            )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_progress_excel(class_id: int) -> bytes:
    """تصدير تقدم الطلاب إلى Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "التقدم"
    ws.append(["الاسم", "البريد الإلكتروني", "الدرس", "الحالة", "الوقت (دقيقة)", "النسبة %"])

    progress = StudentProgress.query.filter_by(class_id=class_id).all()
    for p in progress:
        ws.append(
            [
                p.student.name_ar or "",
                p.student.email,
                p.lesson.title if p.lesson else f"درس #{p.lesson_id}",
                {"not_started": "لم يبدأ", "in_progress": "قيد التنفيذ", "completed": "مكتمل"}[p.status],
                p.seconds_spent // 60,
                p.progress_pct,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
